from openpyxl import load_workbook
wb = load_workbook(filename='product.xlsx', read_only=True)
ws = wb['Sheet1']

text = "ค้นหา W45689" 
if (text.startswith("ค้นหา")): 
    text_id = text.strip("ค้นหา").replace(" ", "")
    # "ค้นหา W45689" --> " W45689" --> "W45689"
    print(text_id)
    
for i in range(2,ws.max_row+1):              # เริ่มอ่านที่ cell A2 เป็นต้นไป
    if text_id == str(ws["A"+str(i)].value): # ค้นหารหัสสินค้า
        product_name = str(ws["B"+str(i)].value) # ชื่อสินค้า
        product_price = str(ws["C"+str(i)].value) # ราคา
        text_out = "รหัสสินค้า " + text_id + " " + product_name + " ราคา " + product_price + " บาท"
        break
    else:
        text_out = "ไม่พบรหัสสินค้านี้"
print(text_out)   

wb.close()
