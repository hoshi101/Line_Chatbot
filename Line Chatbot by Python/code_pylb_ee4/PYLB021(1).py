from openpyxl import load_workbook
from flask import Flask, request
from linebot import LineBotApi, WebhookHandler
from linebot.models import (MessageEvent,
                            TextMessage,
                            TextSendMessage)

from openpyxl import load_workbook

channel_secret = "xxxxxxxx"
channel_access_token = "xxxxxxxxxxxxxxxxx"

line_bot_api = LineBotApi(channel_access_token)
handler = WebhookHandler(channel_secret)

app = Flask(__name__)

@app.route("/", methods=["GET","POST"])
def home():
    try:
        signature = request.headers["X-Line-Signature"]
        body = request.get_data(as_text=True)
        handler.handle(body, signature)
    except:
        pass
    
    return "Hello Line Chatbot"

@handler.add(MessageEvent, message=TextMessage)
def handle_text_message(event):
    text = event.message.text
    print(text)

    wb = load_workbook(filename='star_product.xlsx', read_only=True)
    ws = wb['Sheet1']

    if text.startswith("ค้นหา"):
        text_id = text.strip("ค้นหา").strip()  # ลบคำว่า "ค้นหา" และช่องว่างข้างหน้าและข้างหลัง
        print(text_id)
        
        found = False
        for i in range(2, ws.max_row + 1):
            star_name = str(ws["A" + str(i)].value).strip()  # อ่านและลบช่องว่างในชื่อดวงดาว
            if text_id == star_name:
                star_history = str(ws["B" + str(i)].value).strip()  # อ่านและลบช่องว่างในประวัติ
                text_out = f"ดวงดาว {star_name} {star_history}"
                found = True
                break

        if not found:
            text_out = "ไม่พบดวงดาวนี้"
        
        wb.close()
        
        line_bot_api.reply_message(event.reply_token,
                                   TextSendMessage(text=text_out))

if __name__ == "__main__":          
    app.run()

