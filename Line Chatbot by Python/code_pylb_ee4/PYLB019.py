from openpyxl import load_workbook
wb = load_workbook(filename='product.xlsx', read_only=True)
ws = wb['Sheet1']

for i in range(2,ws.max_row+1):      # เริ่มอ่านที่ cell A2 เป็นต้นไป
    print(str(ws["A"+str(i)].value)) # รหัสสินค้า "A2"
    print(str(ws["B"+str(i)].value)) # ชื่อสินค้า
    print(str(ws["C"+str(i)].value)) # ราคา
    print("-----------------")

wb.close()
